try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
    TKINTER_AVAILABLE = True
except ImportError:
    TKINTER_AVAILABLE = False
    print("Warning: tkinter is not available. GUI functionality will be disabled.")

import os
from pathlib import Path


class GUIHandler:
    def __init__(self):
        """Initialize the GUI Handler with default values."""
        self.input_filename = ""
        self.sheet_name = ""
        self.output_filename = ""
        self.website_url = ""
        self.uni1_name = ""
        self.uni2_name = ""
        self.user_confirmed = False
        
        if not TKINTER_AVAILABLE:
            raise ImportError("tkinter is not available. Cannot create GUI.")
        
        try:
            # Create the main window
            self.root = tk.Tk()
            self.root.title("Country Risk Scraper - Configuration")
            self.root.geometry("700x500")
            self.root.resizable(True, True)
            
            # Variables for form inputs
            self.input_file_var = tk.StringVar()
            self.sheet_name_var = tk.StringVar()
            self.output_file_var = tk.StringVar(value="country.xlsx")
            self.website_url_var = tk.StringVar(value="https://immi.homeaffairs.gov.au/visas/web-evidentiary-tool")
            self.uni1_var = tk.StringVar()
            self.uni2_var = tk.StringVar()
            
            self.setup_gui()
        except Exception as e:
            print(f"Error initializing GUI: {e}")
            raise
    
    def setup_gui(self):
        """Set up the GUI components."""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights for responsive design
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Country Risk Scraper Configuration", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # Input Excel File Section
        ttk.Label(main_frame, text="Input Excel File:", 
                 font=("Arial", 10, "bold")).grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        
        input_frame = ttk.Frame(main_frame)
        input_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        input_frame.columnconfigure(0, weight=1)
        
        self.input_file_entry = ttk.Entry(input_frame, textvariable=self.input_file_var, 
                                         width=50, state="readonly")
        self.input_file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Button(input_frame, text="Browse", 
                  command=self.browse_input_file).grid(row=0, column=1)
        
        # Sheet Name Section
        ttk.Label(main_frame, text="Sheet Name:", 
                 font=("Arial", 10, "bold")).grid(row=3, column=0, sticky=tk.W, pady=(0, 5))
        
        sheet_frame = ttk.Frame(main_frame)
        sheet_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        sheet_frame.columnconfigure(0, weight=1)
        
        self.sheet_name_entry = ttk.Entry(sheet_frame, textvariable=self.sheet_name_var, width=50)
        self.sheet_name_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Button(sheet_frame, text="Auto-detect", 
                  command=self.auto_detect_sheets).grid(row=0, column=1)
        
        # Sheet selection combobox (initially hidden)
        self.sheet_combo_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_combo_var, 
                                       state="readonly", width=47)
        self.sheet_combo.bind('<<ComboboxSelected>>', self.on_sheet_selected)
        
        # Website URL Section
        ttk.Label(main_frame, text="Document Checklist Website:", 
                 font=("Arial", 10, "bold")).grid(row=5, column=0, sticky=tk.W, pady=(0, 5))
        
        website_frame = ttk.Frame(main_frame)
        website_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        website_frame.columnconfigure(0, weight=1)
        
        self.website_entry = ttk.Entry(website_frame, textvariable=self.website_url_var, width=50)
        self.website_entry.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # University Names Section
        ttk.Label(main_frame, text="University Names:", 
                 font=("Arial", 10, "bold")).grid(row=7, column=0, sticky=tk.W, pady=(15, 5))
        
        uni_frame = ttk.Frame(main_frame)
        uni_frame.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        uni_frame.columnconfigure(1, weight=1)
        uni_frame.columnconfigure(3, weight=1)
        
        ttk.Label(uni_frame, text="University 1:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.uni1_entry = ttk.Entry(uni_frame, textvariable=self.uni1_var, width=25)
        self.uni1_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 20))
        
        ttk.Label(uni_frame, text="University 2:").grid(row=0, column=2, sticky=tk.W, padx=(0, 10))
        self.uni2_entry = ttk.Entry(uni_frame, textvariable=self.uni2_var, width=25)
        self.uni2_entry.grid(row=0, column=3, sticky=(tk.W, tk.E))
        
        # Output File Section
        ttk.Label(main_frame, text="Output Excel File:", 
                 font=("Arial", 10, "bold")).grid(row=9, column=0, sticky=tk.W, pady=(15, 5))
        
        output_frame = ttk.Frame(main_frame)
        output_frame.grid(row=10, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 20))
        output_frame.columnconfigure(0, weight=1)
        
        self.output_file_entry = ttk.Entry(output_frame, textvariable=self.output_file_var, width=50)
        self.output_file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Button(output_frame, text="Browse", 
                  command=self.browse_output_file).grid(row=0, column=1)
        
        # Action Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=11, column=0, columnspan=3, pady=(20, 0))
        
        ttk.Button(button_frame, text="Start Scraping", 
                  command=self.start_scraping, 
                  style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(button_frame, text="Cancel", 
                  command=self.cancel).pack(side=tk.LEFT)
        
        # Status label
        self.status_label = ttk.Label(main_frame, text="Please configure the settings above", 
                                     foreground="blue")
        self.status_label.grid(row=12, column=0, columnspan=3, pady=(20, 0))
        
    def browse_input_file(self):
        """Open file dialog to select input Excel file."""
        filename = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=os.getcwd()
        )
        if filename:
            self.input_file_var.set(filename)
            self.status_label.config(text="Input file selected successfully", foreground="green")
    
    def browse_output_file(self):
        """Open file dialog to select output Excel file location."""
        filename = filedialog.asksaveasfilename(
            title="Save Output Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=os.getcwd()
        )
        if filename:
            self.output_file_var.set(filename)
    
    def auto_detect_sheets(self):
        """Auto-detect available sheets in the selected Excel file."""
        if not self.input_file_var.get():
            messagebox.showwarning("Warning", "Please select an input Excel file first!")
            return
        
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=self.input_file_var.get())
            sheet_names = wb.sheetnames
            wb.close()
            
            if sheet_names:
                # Show combobox with available sheets
                self.sheet_combo['values'] = sheet_names
                self.sheet_combo.grid(row=1, column=0, sticky=(tk.W, tk.E), padx=(0, 10), pady=(5, 0))
                
                # If there's only one sheet, select it automatically
                if len(sheet_names) == 1:
                    self.sheet_combo_var.set(sheet_names[0])
                    self.sheet_name_var.set(sheet_names[0])
                    self.status_label.config(text=f"Auto-detected sheet: {sheet_names[0]}", 
                                           foreground="green")
                else:
                    self.status_label.config(text="Multiple sheets found. Please select one from dropdown.", 
                                           foreground="blue")
            else:
                messagebox.showinfo("Info", "No sheets found in the selected file.")
                
        except Exception as e:
            messagebox.showerror("Error", f"Error reading Excel file: {str(e)}")
            self.status_label.config(text="Error reading Excel file", foreground="red")
    
    def on_sheet_selected(self, event=None):
        """Handle sheet selection from combobox."""
        selected_sheet = self.sheet_combo_var.get()
        self.sheet_name_var.set(selected_sheet)
        self.status_label.config(text=f"Sheet selected: {selected_sheet}", foreground="green")
    
    def validate_inputs(self):
        """Validate all required inputs."""
        if not self.input_file_var.get():
            messagebox.showwarning("Warning", "Please select an input Excel file!")
            return False
        
        if not os.path.exists(self.input_file_var.get()):
            messagebox.showerror("Error", "The selected input file does not exist!")
            return False
        
        if not self.sheet_name_var.get():
            messagebox.showwarning("Warning", "Please enter or select a sheet name!")
            return False
        
        if not self.website_url_var.get():
            messagebox.showwarning("Warning", "Please enter the document checklist website URL!")
            return False
        
        if not self.uni1_var.get():
            messagebox.showwarning("Warning", "Please enter the name for University 1!")
            return False
        
        if not self.uni2_var.get():
            messagebox.showwarning("Warning", "Please enter the name for University 2!")
            return False
        
        if not self.output_file_var.get():
            messagebox.showwarning("Warning", "Please specify an output file name!")
            return False
        
        # Validate URL format (basic check)
        website_url = self.website_url_var.get().strip()
        if not (website_url.startswith('http://') or website_url.startswith('https://')):
            messagebox.showwarning("Warning", "Website URL should start with http:// or https://")
            return False
        
        # Check if output directory exists
        output_dir = os.path.dirname(self.output_file_var.get())
        if output_dir and not os.path.exists(output_dir):
            try:
                os.makedirs(output_dir, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", f"Cannot create output directory: {str(e)}")
                return False
        
        return True
    
    def start_scraping(self):
        """Start the scraping process after validation."""
        if self.validate_inputs():
            self.input_filename = self.input_file_var.get()
            self.sheet_name = self.sheet_name_var.get()
            self.output_filename = self.output_file_var.get()
            self.website_url = self.website_url_var.get().strip()
            self.uni1_name = self.uni1_var.get().strip()
            self.uni2_name = self.uni2_var.get().strip()
            self.user_confirmed = True
            
            self.status_label.config(text="Configuration confirmed! Starting scraper...", 
                                   foreground="green")
            self.root.quit()  # Exit the GUI event loop
    
    def cancel(self):
        """Cancel the operation and close the GUI."""
        self.user_confirmed = False
        self.root.quit()
    
    def get_user_inputs(self):
        """Show the GUI and return user inputs when confirmed."""
        # Center the window
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        
        # Show the GUI and wait for user interaction
        self.root.mainloop()
        
        # Clean up
        self.root.destroy()
        
        if self.user_confirmed:
            return {
                'input_filename': self.input_filename,
                'sheet_name': self.sheet_name,
                'output_filename': self.output_filename,
                'website_url': self.website_url,
                'uni1_name': self.uni1_name,
                'uni2_name': self.uni2_name
            }
        else:
            return None
    
    @staticmethod
    def run_gui():
        """Static method to run the GUI and get user inputs."""
        if not TKINTER_AVAILABLE:
            return GUIHandler.run_console_interface()
        
        try:
            gui = GUIHandler()
            return gui.get_user_inputs()
        except Exception as e:
            print(f"GUI Error: {e}")
            print("Falling back to console interface...")
            return GUIHandler.run_console_interface()
    
    @staticmethod
    def run_console_interface():
        """Fallback console interface when GUI is not available."""
        print("\n" + "="*60)
        print("COUNTRY RISK SCRAPER - CONSOLE CONFIGURATION")
        print("="*60)
        print("GUI is not available. Using console interface instead.\n")
        
        try:
            # Get input file
            while True:
                input_file = input("Enter the path to your input Excel file: ").strip()
                if input_file and os.path.exists(input_file):
                    break
                elif input_file and not os.path.exists(input_file):
                    print(f"Error: File '{input_file}' does not exist. Please try again.")
                else:
                    print("Please enter a valid file path.")
            
            # Get sheet name with auto-detection
            sheet_name = ""
            try:
                from openpyxl import load_workbook
                wb = load_workbook(filename=input_file)
                sheet_names = wb.sheetnames
                wb.close()
                
                if len(sheet_names) == 1:
                    sheet_name = sheet_names[0]
                    print(f"Auto-detected sheet: {sheet_name}")
                elif len(sheet_names) > 1:
                    print(f"\nAvailable sheets: {', '.join(sheet_names)}")
                    while True:
                        sheet_input = input(f"Enter sheet name (or press Enter for '{sheet_names[0]}'): ").strip()
                        if not sheet_input:
                            sheet_name = sheet_names[0]
                            break
                        elif sheet_input in sheet_names:
                            sheet_name = sheet_input
                            break
                        else:
                            print(f"Error: '{sheet_input}' is not a valid sheet name.")
                else:
                    print("Warning: No sheets found in the file.")
                    sheet_name = input("Enter sheet name manually: ").strip()
                    
            except Exception as e:
                print(f"Warning: Could not read Excel file sheets ({e})")
                sheet_name = input("Enter sheet name: ").strip()
            
            if not sheet_name:
                print("Error: Sheet name is required.")
                return None
            
            # Get website URL
            default_url = "https://immi.homeaffairs.gov.au/visas/web-evidentiary-tool"
            website_url = input(f"Enter document checklist website URL (or press Enter for default): ").strip()
            if not website_url:
                website_url = default_url
            
            # Get university names
            uni1_name = ""
            while not uni1_name:
                uni1_name = input("Enter University 1 name: ").strip()
                if not uni1_name:
                    print("University 1 name is required.")
            
            uni2_name = ""
            while not uni2_name:
                uni2_name = input("Enter University 2 name: ").strip()
                if not uni2_name:
                    print("University 2 name is required.")
            
            # Get output file
            default_output = "country.xlsx"
            output_file = input(f"Enter output file name (or press Enter for '{default_output}'): ").strip()
            if not output_file:
                output_file = default_output
            
            # Confirm settings
            print("\n" + "="*60)
            print("CONFIGURATION SUMMARY:")
            print("="*60)
            print(f"Input file:     {input_file}")
            print(f"Sheet name:     {sheet_name}")
            print(f"Website URL:    {website_url}")
            print(f"University 1:   {uni1_name}")
            print(f"University 2:   {uni2_name}")
            print(f"Output file:    {output_file}")
            print("="*60)
            
            while True:
                confirm = input("Continue with these settings? (y/n): ").strip().lower()
                if confirm in ['y', 'yes']:
                    return {
                        'input_filename': input_file,
                        'sheet_name': sheet_name,
                        'output_filename': output_file,
                        'website_url': website_url,
                        'uni1_name': uni1_name,
                        'uni2_name': uni2_name
                    }
                elif confirm in ['n', 'no']:
                    print("Configuration cancelled.")
                    return None
                else:
                    print("Please enter 'y' for yes or 'n' for no.")
                    
        except KeyboardInterrupt:
            print("\nOperation cancelled by user.")
            return None
        except Exception as e:
            print(f"Error in console interface: {e}")
            return None


# Example usage
if __name__ == "__main__":
    # Test the GUI/Console interface
    result = GUIHandler.run_gui()
    if result:
        print(f"\nUser selected:")
        print(f"  Input file: {result['input_filename']}")
        print(f"  Sheet name: {result['sheet_name']}")
        print(f"  Website URL: {result['website_url']}")
        print(f"  University 1: {result['uni1_name']}")
        print(f"  University 2: {result['uni2_name']}")
        print(f"  Output file: {result['output_filename']}")
    else:
        print("User cancelled the operation.")