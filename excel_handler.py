from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv
import os

class ExcelHandler:
    def __init__(self, filename, sheetname, uni1_name=None, uni2_name=None):
        """Initialize the Excel handler by loading the workbook and extracting country names."""
        load_dotenv()
        # Use provided university names or fallback to environment variables
        self.uni1 = uni1_name if uni1_name else os.getenv("UNI1")
        self.uni2 = uni2_name if uni2_name else os.getenv("UNI2")
        
        self.filename = filename
        self.wb = load_workbook(filename=self.filename)
        self.sheet = self.wb[sheetname]

        # Read country names from Column A (Rows 3 to 238)
        self.countries = [self.sheet[f"A{row}"].value for row in range(3, 238)]
        
        # Initialize nested dictionary with default values using UNI1 and UNI2 keys
        self.country_data = {country: {self.uni1: "N", self.uni2: "N"} for country in self.countries}

    def update_country_data(self, country, uni1_value, uni2_value):
        """Update the UNI1 and UNI2 values for a specific country."""
        if country in self.country_data:
            self.country_data[country][self.uni1] = uni1_value
            self.country_data[country][self.uni2] = uni2_value
        else:
            print(f"Warning: {country} not found in Excel list!")

    def save_to_excel(self):
        """Write updated UNI1 and UNI2 values back to the original Excel file and save."""
        for i, country in enumerate(self.countries, start=3):  # Row 3 to 238
            self.sheet[f"B{i}"] = self.country_data[country][self.uni1]
            self.sheet[f"C{i}"] = self.country_data[country][self.uni2]

        self.wb.save(self.filename)
        print(f"Updated data saved to {self.filename}")

    def export_to_excel(self, output_filename="country.xlsx"):
        """Exports the country data dictionary to a new Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Country Data"

        # Write headers using UNI1 and UNI2 from .env
        ws.append(["Country", self.uni1, self.uni2])

        # Write data
        for country, values in self.country_data.items():
            ws.append([country, values[self.uni1], values[self.uni2]])

        wb.save(output_filename)
        print(f"Data successfully exported to {output_filename}")
